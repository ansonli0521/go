import openpyxl
from elo.models import Player, Game
from datetime import date, timedelta
from django.db.models import Q
import traceback
import math
from elo.constants import source, out_source

# Function to calculate the Probability
def probability(rating1, rating2):

    return 1.0 * 1.0 / (1 + 1.0 * math.pow(10, 1.0 * (rating1 - rating2) / 400))


# Function to calculate Elo rating
# K is a constant.
# d determines whether
# Player A wins or Player B.
def elo_rating(Ra, Rb, d, new, k, f):

    # To calculate the Winning
    # Probability of Player B
    Pb = probability(Ra, Rb)

    # To calculate the Winning
    # Probability of Player A
    Pa = probability(Rb, Ra)

    K = k

    if new == 'Z':
    
        if (d == 'B'):
            Ra = Ra + K * (1 - Pa)
            Rb = Rb + K * (0 - Pb)

        elif (d == 'W'):
            Ra = Ra + K * (0 - Pa)
            Rb = Rb + K * (1 - Pb)
        
        else:
            Ra = Ra + K * (0.5 - Pa)
            Rb = Rb + K * (0.5 - Pb)
    
    if new == 'B':

        if (d == 'B'):
            Ra = Ra + f * K * (1 - Pa)
            Rb = Rb + K * (0 - Pb)

        elif (d == 'W'):
            Ra = Ra + K * (0 - Pa)
            Rb = Rb + K * (1 - Pb)
        
        else:
            Ra = Ra + K * (0.5 - Pa)
            Rb = Rb + K * (0.5 - Pb)
    
    if new == 'W':

        if (d == 'B'):
            Ra = Ra + K * (1 - Pa)
            Rb = Rb + K * (0 - Pb)

        elif (d == 'W'):
            Ra = Ra + K * (0 - Pa)
            Rb = Rb + f * K * (1 - Pb)
        
        else:
            Ra = Ra + K * (0.5 - Pa)
            Rb = Rb + K * (0.5 - Pb)

    if new == 'T':

        if (d == 'B'):
            Ra = Ra + f * K * (1 - Pa)
            Rb = Rb + K * (0 - Pb)

        elif (d == 'W'):
            Ra = Ra + K * (0 - Pa)
            Rb = Rb + f * K * (1 - Pb)
        
        else:
            Ra = Ra + K * (0.5 - Pa)
            Rb = Rb + K * (0.5 - Pb)

    return Ra, Rb

def calculate(k,f):
    try:
        players = Player.objects.all()
        games = Game.objects.order_by('game_date')
        start_date = date(2022, 5, 11)
        book = openpyxl.load_workbook(source)
        history_sheet = book.create_sheet("History")
        history_sheet.cell(row=1, column=1).value = 'Date'
        history_sheet.cell(row=1, column=2).value = 'Player'
        history_sheet.cell(row=1, column=3).value = 'Opponent'
        history_sheet.cell(row=1, column=4).value = 'Old Elo'
        history_sheet.cell(row=1, column=5).value = 'New Elo'
        history_sheet.cell(row=1, column=6).value = 'Elo Change'
        c=2
        for game in games:
            black_player = game.black
            white_player = game.white
            black_old_elo = black_player.elo
            white_old_elo = white_player.elo
            if black_player.total_games >= 5 and white_player.total_games >= 5:
                black_player.elo, white_player.elo = elo_rating(black_old_elo, white_old_elo, game.result, 'Z', k, f)
            elif black_player.total_games < 5 and white_player.total_games >= 5:
                black_player.elo, white_player.elo = elo_rating(black_old_elo, white_old_elo, game.result, 'B', k, f)
            elif black_player.total_games >= 5 and white_player.total_games < 5:
                black_player.elo, white_player.elo = elo_rating(black_old_elo, white_old_elo, game.result, 'W', k, f)
            elif black_player.total_games < 5 and white_player.total_games < 5:
                black_player.elo, white_player.elo = elo_rating(black_old_elo, white_old_elo, game.result, 'T', k, f)
            if game.result == 'B':
                black_player.win += 1
                white_player.lost += 1
            elif game.result == 'W':
                black_player.lost += 1
                white_player.win += 1
            black_player.total_games += 1
            white_player.total_games += 1
            black_player.save()
            white_player.save()
            game.black_old_elo = black_old_elo
            game.white_old_elo = white_old_elo
            game.black_new_elo = black_player.elo
            game.white_new_elo = white_player.elo
            game.save()
            history_sheet.cell(row=c, column=1).value = game.game_date
            history_sheet.cell(row=c, column=2).value = black_player.name
            history_sheet.cell(row=c, column=3).value = white_player.name
            history_sheet.cell(row=c, column=4).value = black_old_elo
            history_sheet.cell(row=c, column=5).value = black_player.elo
            history_sheet.cell(row=c, column=6).value = black_player.elo - black_old_elo
            history_sheet.cell(row=c+1, column=1).value = game.game_date
            history_sheet.cell(row=c+1, column=2).value = white_player.name
            history_sheet.cell(row=c+1, column=3).value = black_player.name
            history_sheet.cell(row=c+1, column=4).value = white_old_elo
            history_sheet.cell(row=c+1, column=5).value = white_player.elo
            history_sheet.cell(row=c+1, column=6).value = white_player.elo - white_old_elo
            c+=2

        elo_sheet = book.create_sheet('Elo')
        elo_sheet.cell(row=1, column=1).value = 'Player'
        elo_sheet.cell(row=1, column=2).value = 'Elo'
        elo_sheet.cell(row=1, column=3).value = 'Total Games'
        elo_sheet.cell(row=1, column=4).value = 'Win'
        elo_sheet.cell(row=1, column=5).value = 'Lost'
        elo_sheet.cell(row=1, column=6).value = 'Win Rate'
        elo_sheet.cell(row=1, column=7).value = 'Status'
        c=2
        players = Player.objects.order_by('-elo')
        for player in players:
            latest_game = Game.objects.filter(Q(black=player) | Q(white=player)).order_by('-game_date').first()
            if latest_game.game_date + timedelta(days=180) < date.today():
                player.status = 'inactive'
            elif player.total_games <= 5:
                player.status = 'new'
            else:
                player.status = 'normal'
            player.winrate = str(round((player.win / player.total_games) * 10000) / 100) + '%'
            player.save()

            elo_sheet.cell(row=c, column=1).value = player.name
            elo_sheet.cell(row=c, column=2).value = player.elo
            elo_sheet.cell(row=c, column=3).value = player.total_games
            elo_sheet.cell(row=c, column=4).value = player.win
            elo_sheet.cell(row=c, column=5).value = player.lost
            elo_sheet.cell(row=c, column=6).value = player.winrate
            elo_sheet.cell(row=c, column=7).value = player.status
            c+=1

        book.save(out_source)
        book.close()
    except:
        book.close()
        traceback.print_exc()