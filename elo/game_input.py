import openpyxl
from elo.models import Player, Game
import traceback
from elo.constants import source, rank_elo_map


def delete_all_games():
    games = Game.objects.all()
    players = Player.objects.all()
    games.delete()
    players.delete()
    
def game_input():
    try:
        book = openpyxl.load_workbook(source)
        player_sheet = book.get_sheet_by_name('Player')
        c = 2
        player_name = player_sheet.cell(row=c, column=2).value
        ini_rank = player_sheet.cell(row=c, column=3).value
        while player_name and ini_rank is not None:
            ini_elo = rank_elo_map[ini_rank]
            new_player = Player(name=player_name, elo=ini_elo, total_games=0)
            new_player.save()
            c += 1
            player_name = player_sheet.cell(row=c, column=2).value
            ini_rank = player_sheet.cell(row=c, column=3).value
        print('player data input finished')

        game_sheet = book.get_sheet_by_name('Record')
        c = 2
        black = game_sheet.cell(row=c, column=3).value
        white = game_sheet.cell(row=c, column=5).value
        handicap = game_sheet.cell(row=c, column=8).value
        result = game_sheet.cell(row=c, column=10).value
        game_date = game_sheet.cell(row=c, column=2).value
        while black is not None and white is not None and result is not None and game_date is not None:
            black_player = Player.objects.get(name=black)
            white_player = Player.objects.get(name=white)
            new_game = Game(black=black_player, white=white_player, handicap=handicap, result=result, game_date=game_date)
            new_game.save()
            c += 1
            black = game_sheet.cell(row=c, column=3).value
            white = game_sheet.cell(row=c, column=5).value
            handicap = game_sheet.cell(row=c, column=8).value
            result = game_sheet.cell(row=c, column=10).value
            game_date = game_sheet.cell(row=c, column=2).value
        print('game data input finished')
        book.close()
    except:
        book.close()
        traceback.print_exc()