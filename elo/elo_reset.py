import openpyxl
from elo.models import Player
import traceback

rank_elo_map = {
    '12k': 900,
    '8k': 1300,
    '4k': 1700,
    '1d': 2100,
    '3d': 2300,
    '5d': 2500
}

def reset():
    try:
        book = openpyxl.load_workbook('/Users/ansonli/go/elo/LIHKG-Record.xlsx')
        player_sheet = book.get_sheet_by_name('Player')
        c = 2
        player_name = player_sheet.cell(row=c, column=1).value
        ini_rank = player_sheet.cell(row=c, column=2).value
        while player_name and ini_rank is not None:
            ini_elo = rank_elo_map[ini_rank]
            player = Player.objects.get(name=player_name)
            player.elo = ini_elo
            player.total_games = 0
            player.save()
            c += 1
            player_name = player_sheet.cell(row=c, column=1).value
            ini_rank = player_sheet.cell(row=c, column=2).value
        print('player elo reset finished')
    except:
        book.close()
        traceback.print_exc()