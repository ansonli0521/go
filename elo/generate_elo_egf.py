import openpyxl
from elo.models import Player, Game
from datetime import date, timedelta
from django.db.models import Q
import traceback
from decimal import Decimal

"""
EGF Official Rating System (since 2021)
https://www.europeangodatabase.eu/EGD/EGF_rating_system.php
"""


def con(r: Decimal) -> Decimal:
    """Rating volatility factor"""
    return ((3300 - r) / 200) ** Decimal("1.6")


def beta(r: Decimal) -> Decimal:
    """Mapping function for EGD ratings"""
    return -7 * (3300 - r).ln()


def bonus(r: Decimal) -> Decimal:
    """Term to counter rating deflation"""
    return (((2300 - r) / 80).exp() + 1).ln() / 5


def win_prob(r1: Decimal, r2: Decimal) -> Decimal:
    """Expected result computed by Bradley-Terry formula"""
    return 1 / (1 + (beta(r2) - beta(r1)).exp())


def calculate(
    r1: float,
    r2: float,
    sa: float,
    color: str,
    handicap: int,
    k,
) -> Decimal:
    """Calculate new rating of the player
    :param r1: EGD rating of the player
    :param r2: EGD rating of the opponent
    :param sa: actual game result, either 0, 0.5 or 1
    :param color: either "white" or "black", has no effect without handicap
    :param handicap: between 0 and 9
    """
    r1 = r1_adj = Decimal(r1)
    r2_adj = Decimal(r2)
    sa = Decimal(sa)
    
    assert sa in (0, 0.5, 1), "sa should be either 0, 0.5 or 1"
    assert color in ("white", "black"), "color should be either 'white' or 'black'"
    assert 0 <= handicap <= 9, "handicap should be between 0 and 9"

    if handicap > 0:
        if color == "white":
            r2_adj += 50 + 100 * (handicap - 1)
        else:
            r1_adj += 50 + 100 * (handicap - 1)

    se = win_prob(r1_adj, r2_adj)

    # return r1 + con(r1) * (sa - se) + bonus(r1)
    return r1 + k * (con(r1) * (sa - se))

def first_five_calculate(
    r1: float,
    r2: float,
    sa: float,
    color,
    handicap,
    k,
    f,
) -> Decimal:
    """Calculate new rating of the player
    :param r1: EGD rating of the player
    :param r2: EGD rating of the opponent
    :param sa: actual game result, either 0, 0.5 or 1
    :param color: either "white" or "black", has no effect without handicap
    :param handicap: between 0 and 9
    """
    r1 = r1_adj = Decimal(r1)
    r2_adj = Decimal(r2)
    sa = Decimal(sa)
    
    assert sa in (0, 0.5, 1), "sa should be either 0, 0.5 or 1"
    assert color in ("white", "black"), "color should be either 'white' or 'black'"
    assert 0 <= handicap <= 9, "handicap should be between 0 and 9"

    if handicap > 0:
        if color == "white":
            r2_adj += 50 + 100 * (handicap - 1)
        else:
            r1_adj += 50 + 100 * (handicap - 1)

    se = win_prob(r1_adj, r2_adj)

    # first five games double elo change
    return r1 + f * k * (con(r1) * (sa - se))

def egf_calculate(k,f):
    try:
        players = Player.objects.all()
        games = Game.objects.order_by('game_date')
        start_date = date(2022, 5, 11)
        book = openpyxl.load_workbook('/Users/ansonli/go/elo/LIHKG-Record.xlsx')
        history_sheet = book.create_sheet("History")
        history_sheet.cell(row=1, column=1).value = 'Date'
        history_sheet.cell(row=1, column=2).value = 'Player'
        history_sheet.cell(row=1, column=3).value = 'Opponent'
        history_sheet.cell(row=1, column=4).value = 'Old Elo'
        history_sheet.cell(row=1, column=5).value = 'New Elo'
        history_sheet.cell(row=1, column=6).value = 'Elo Change'
        c=2
        for game in games:
            black_player = game.black
            white_player = game.white
            black_old_elo = black_player.elo
            white_old_elo = white_player.elo
            if game.result == 'B':
                if black_player.total_games < 5:
                    black_player.elo = first_five_calculate(black_old_elo, white_old_elo, 1, 'black', game.handicap, k, f)
                else:
                    black_player.elo = calculate(black_old_elo, white_old_elo, 1, 'black', game.handicap, k)
                # if white_player.total_games < 5:
                #     white_player.elo = first_five_calculate(white_old_elo, black_old_elo, 0, 'white', game.handicap)
                # else:
                #     white_player.elo = calculate(white_old_elo, black_old_elo, 0, 'white', game.handicap)
                white_player.elo = calculate(white_old_elo, black_old_elo, 0, 'white', game.handicap, k)
            elif game.result == 'W':
                # if black_player.total_games < 5:
                #     black_player.elo = first_five_calculate(black_old_elo, white_old_elo, 0, 'black', game.handicap)
                # else:
                #     black_player.elo = calculate(black_old_elo, white_old_elo, 0, 'black', game.handicap)
                black_player.elo = calculate(black_old_elo, white_old_elo, 0, 'black', game.handicap, k)
                if white_player.total_games < 5:
                    white_player.elo = first_five_calculate(white_old_elo, black_old_elo, 1, 'white', game.handicap, k, f)
                else:
                    white_player.elo = calculate(white_old_elo, black_old_elo, 1, 'white', game.handicap, k)        
            elif game.result == 'D':
                if black_player.total_games < 5:
                    black_player.elo = first_five_calculate(black_old_elo, white_old_elo, 0.5, 'black', game.handicap, k, f)
                else:
                    black_player.elo = calculate(black_old_elo, white_old_elo, 0.5, 'black', game.handicap, k)
                if white_player.total_games < 5:
                    white_player.elo = first_five_calculate(white_old_elo, black_old_elo, 0.5, 'white', game.handicap, k, f)
                else:
                    white_player.elo = calculate(white_old_elo, black_old_elo, 0.5, 'white', game.handicap, k)
            black_player.total_games += 1
            white_player.total_games += 1
            black_player.save()
            white_player.save()
            game.black_old_elo = black_old_elo
            game.white_old_elo = white_old_elo
            game.black_new_elo = black_player.elo
            game.white_new_elo = white_player.elo
            game.save()
            history_sheet.cell(row=c, column=1).value = game.game_date
            history_sheet.cell(row=c, column=2).value = black_player.name
            history_sheet.cell(row=c, column=3).value = white_player.name
            history_sheet.cell(row=c, column=4).value = black_old_elo
            history_sheet.cell(row=c, column=5).value = black_player.elo
            history_sheet.cell(row=c, column=6).value = black_player.elo - Decimal(black_old_elo)
            history_sheet.cell(row=c+1, column=1).value = game.game_date
            history_sheet.cell(row=c+1, column=2).value = white_player.name
            history_sheet.cell(row=c+1, column=3).value = black_player.name
            history_sheet.cell(row=c+1, column=4).value = white_old_elo
            history_sheet.cell(row=c+1, column=5).value = white_player.elo
            history_sheet.cell(row=c+1, column=6).value = white_player.elo - Decimal(white_old_elo)
            c+=2

        elo_sheet = book.create_sheet('Elo')
        elo_sheet.cell(row=1, column=1).value = 'Player'
        elo_sheet.cell(row=1, column=2).value = 'Elo'
        elo_sheet.cell(row=1, column=3).value = 'Total Games'
        elo_sheet.cell(row=1, column=4).value = 'Status'
        c=2
        players = Player.objects.order_by('-elo')
        for player in players:
            latest_game = Game.objects.filter(Q(black=player) | Q(white=player)).order_by('-game_date').first()
            if latest_game.game_date + timedelta(days=180) < date.today():
                player.status = 'inactive'
            elif player.total_games <= 5:
                player.status = 'new'
            else:
                player.status = 'normal'
            player.save()

            elo_sheet.cell(row=c, column=1).value = player.name
            elo_sheet.cell(row=c, column=2).value = player.elo
            elo_sheet.cell(row=c, column=3).value = player.total_games
            elo_sheet.cell(row=c, column=4).value = player.status
            c+=1

        book.save('/Users/ansonli/go/elo/LIHKG-Record-with-Elo.xlsx')
        book.close()
    except:
        book.close()
        traceback.print_exc()