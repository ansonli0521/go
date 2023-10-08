import openpyxl
from elo.models import Player, Game
from datetime import date, timedelta
import whr
import traceback

try:
    base = whr.Base(config={'w2': 300})
    players = Player.objects.all()
    games = Game.objects.order_by('game_date')
    start_date = date(2022, 5, 11)
    book = openpyxl.load_workbook('LIHKG-Record.xlsx')
    for game in games:
        base.create_game(game.black.name, game.white.name, game.result, (game.game_date - start_date).days)
    base.iterate(100)
    print(base.get_ordered_ratings())

    for player in players:
        print(player.name)
        player.elo = base.ratings_for_player(player.name)[-1][1]
        player.save()

    elo_sheet = book.create_sheet('Elo')
    elo_sheet.cell(row=1, column=1).value = 'Player'
    elo_sheet.cell(row=1, column=2).value = 'Elo'
    elo_sheet.cell(row=1, column=3).value = 'Total Games'
    c=2
    players = Player.objects.order_by('-elo')
    for player in players:
        elo_sheet.cell(row=c, column=1).value = player.name
        elo_sheet.cell(row=c, column=2).value = player.elo
        elo_sheet.cell(row=c, column=3).value = player.total_games
        c+=1
    
    for player in players:
        player_sheet = book.create_sheet(player.name + "'s_elo_history")
        player_sheet.cell(row=1, column=1).value = 'Date'
        player_sheet.cell(row=1, column=2).value = 'Elo'
        c=2
        for rating in base.ratings_for_player(player.name):
            player_sheet.cell(row=c, column=1).value = start_date + timedelta(days=rating[0])
            # player_sheet.cell(row=c, column=2).value = player.elo + rating[1] - base.ratings_for_player(player.name)[0][1]
            player_sheet.cell(row=c, column=2).value = rating[1]
            c += 1

    book.save('LIHKG-Record-with-Elo.xlsx')
    book.close()
except:
    book.close()
    traceback.print_exc()