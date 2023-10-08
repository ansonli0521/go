import openpyxl
from elo.models import Player
import traceback

try:
    book = openpyxl.load_workbook('LIHKG-Record.xlsx')
    elo_sheet = book.create_sheet('Elo')
    elo_sheet.cell(row=1, column=1).value = 'Player'
    elo_sheet.cell(row=1, column=2).value = 'Elo'
    elo_sheet.cell(row=1, column=3).value = 'Total Games'
    c=2
    players = Player.objects.order_by('-elo')
    for player in players:
        elo_sheet.cell(row=c, column=1).value = player.name
        elo_sheet.cell(row=c, column=2).value = player.elo
        elo_sheet.cell(row=c, column=3).value = player.total_games
        c+=1
    book.save('LIHKG-Record-with-Elo.xlsx')
    book.close()
except:
    book.close()
    traceback.print_exc()