import openpyxl
from elo.models import Player, Game
from datetime import date, timedelta
from django.db.models import Q
import traceback
from decimal import Decimal
from elo.constants import source, out_source, csv_out_path
import pandas as pd

"""
EGF Official Rating System (since 2021)
https://www.europeangodatabase.eu/EGD/EGF_rating_system.php
"""


def con(r: Decimal) -> Decimal:
    """Rating volatility factor"""
    return ((3300 - r) / 200) ** Decimal("1.6")


def beta(r: Decimal) -> Decimal:
    """Mapping function for EGD ratings"""
    return -7 * (3300 - r).ln()


def bonus(r: Decimal) -> Decimal:
    """Term to counter rating deflation"""
    return (((2300 - r) / 80).exp() + 1).ln() / 5


def win_prob(r1: Decimal, r2: Decimal) -> Decimal:
    """Expected result computed by Bradley-Terry formula"""
    return 1 / (1 + (beta(r2) - beta(r1)).exp())


def calculate(
    r1: float,
    r2: float,
    sa: float,
    color: str,
    handicap: int,
    k,
) -> Decimal:
    """Calculate new rating of the player
    :param r1: EGD rating of the player
    :param r2: EGD rating of the opponent
    :param sa: actual game result, either 0, 0.5 or 1
    :param color: either "white" or "black", has no effect without handicap
    :param handicap: between 0 and 9
    """
    r1 = r1_adj = Decimal(r1)
    r2_adj = Decimal(r2)
    sa = Decimal(sa)
    
    assert sa in (0, 0.5, 1), "sa should be either 0, 0.5 or 1"
    assert color in ("white", "black"), "color should be either 'white' or 'black'"
    assert 0 <= handicap <= 9, "handicap should be between 0 and 9"

    if handicap > 0:
        if color == "white":
            r2_adj += 50 + 100 * (handicap - 1)
        else:
            r1_adj += 50 + 100 * (handicap - 1)

    se = win_prob(r1_adj, r2_adj)

    # return r1 + con(r1) * (sa - se) + bonus(r1)
    if r1 >= 2100:
        return r1 + 3 * k * (con(r1) * (sa - se))
    elif r1 < 1700:
        return r1 + 2 * k * (con(r1) * (sa - se))
    else:
        return r1 + Decimal(2.5) * k * (con(r1) * (sa - se))

def first_five_calculate(
    r1: float,
    r2: float,
    sa: float,
    color,
    handicap,
    k,
    f,
) -> Decimal:
    """Calculate new rating of the player
    :param r1: EGD rating of the player
    :param r2: EGD rating of the opponent
    :param sa: actual game result, either 0, 0.5 or 1
    :param color: either "white" or "black", has no effect without handicap
    :param handicap: between 0 and 9
    """
    r1 = r1_adj = Decimal(r1)
    r2_adj = Decimal(r2)
    sa = Decimal(sa)
    
    assert sa in (0, 0.5, 1), "sa should be either 0, 0.5 or 1"
    assert color in ("white", "black"), "color should be either 'white' or 'black'"
    assert 0 <= handicap <= 9, "handicap should be between 0 and 9"

    if handicap > 0:
        if color == "white":
            r2_adj += 50 + 100 * (handicap - 1)
        else:
            r1_adj += 50 + 100 * (handicap - 1)

    se = win_prob(r1_adj, r2_adj)

    # first five games double elo change
    if r1 >= 2100:
        return r1 + 3 * f * k * (con(r1) * (sa - se))
    elif r1 < 1700:
        return r1 + 2 * f * k * (con(r1) * (sa - se))
    else:
        return r1 + Decimal(2.5) * f * k * (con(r1) * (sa - se))

def egf_calculate(k,f):
    try:
        players = Player.objects.all()
        games = Game.objects.order_by('game_date')
        start_date = date(2022, 5, 11)
        book = openpyxl.load_workbook(source)
        history_sheet = book.create_sheet("History")
        history_sheet.cell(row=1, column=1).value = 'Date'
        history_sheet.cell(row=1, column=2).value = 'Player'
        history_sheet.cell(row=1, column=3).value = 'Opponent'
        history_sheet.cell(row=1, column=4).value = 'Old Elo'
        history_sheet.cell(row=1, column=5).value = 'New Elo'
        history_sheet.cell(row=1, column=6).value = 'Elo Change'
        dat = []
        ply = []
        opp = []
        o_elo = []
        n_elo = []
        elo_chg = []
        blk = []
        wht = []
        hdcp = []
        rst = []
        gdat = []
        b_o_elo = []
        b_n_elo = []
        w_o_elo = []
        w_n_elo = []
        b_elo_chg = []
        w_elo_chg = []
        c=2
        for game in games:
            black_player = game.black
            white_player = game.white
            black_old_elo = black_player.elo
            white_old_elo = white_player.elo
            if game.result == 'B':
                if black_player.total_games < 5:
                    black_player.elo = first_five_calculate(black_old_elo, white_old_elo, 1, 'black', game.handicap, k, f)
                else:
                    black_player.elo = calculate(black_old_elo, white_old_elo, 1, 'black', game.handicap, k)
                # if white_player.total_games < 5:
                #     white_player.elo = first_five_calculate(white_old_elo, black_old_elo, 0, 'white', game.handicap)
                # else:
                #     white_player.elo = calculate(white_old_elo, black_old_elo, 0, 'white', game.handicap)
                white_player.elo = calculate(white_old_elo, black_old_elo, 0, 'white', game.handicap, k)
                black_player.win += 1
                white_player.lost += 1
            elif game.result == 'W':
                # if black_player.total_games < 5:
                #     black_player.elo = first_five_calculate(black_old_elo, white_old_elo, 0, 'black', game.handicap)
                # else:
                #     black_player.elo = calculate(black_old_elo, white_old_elo, 0, 'black', game.handicap)
                black_player.elo = calculate(black_old_elo, white_old_elo, 0, 'black', game.handicap, k)
                if white_player.total_games < 5:
                    white_player.elo = first_five_calculate(white_old_elo, black_old_elo, 1, 'white', game.handicap, k, f)
                else:
                    white_player.elo = calculate(white_old_elo, black_old_elo, 1, 'white', game.handicap, k)
                black_player.lost += 1
                white_player.win += 1
            elif game.result == 'D':
                if black_player.total_games < 5:
                    black_player.elo = first_five_calculate(black_old_elo, white_old_elo, 0.5, 'black', game.handicap, k, f)
                else:
                    black_player.elo = calculate(black_old_elo, white_old_elo, 0.5, 'black', game.handicap, k)
                if white_player.total_games < 5:
                    white_player.elo = first_five_calculate(white_old_elo, black_old_elo, 0.5, 'white', game.handicap, k, f)
                else:
                    white_player.elo = calculate(white_old_elo, black_old_elo, 0.5, 'white', game.handicap, k)
            black_player.total_games += 1
            white_player.total_games += 1
            black_player.save()
            white_player.save()
            game.black_old_elo = black_old_elo
            game.white_old_elo = white_old_elo
            game.black_new_elo = black_player.elo
            game.white_new_elo = white_player.elo
            game.save()
            history_sheet.cell(row=c, column=1).value = game.game_date
            history_sheet.cell(row=c, column=2).value = black_player.name
            history_sheet.cell(row=c, column=3).value = white_player.name
            history_sheet.cell(row=c, column=4).value = black_old_elo
            history_sheet.cell(row=c, column=5).value = black_player.elo
            history_sheet.cell(row=c, column=6).value = black_player.elo - Decimal(black_old_elo)
            history_sheet.cell(row=c+1, column=1).value = game.game_date
            history_sheet.cell(row=c+1, column=2).value = white_player.name
            history_sheet.cell(row=c+1, column=3).value = black_player.name
            history_sheet.cell(row=c+1, column=4).value = white_old_elo
            history_sheet.cell(row=c+1, column=5).value = white_player.elo
            history_sheet.cell(row=c+1, column=6).value = white_player.elo - Decimal(white_old_elo)
            dat.append(game.game_date)
            ply.append(black_player.name)
            opp.append(white_player.name)
            o_elo.append(black_old_elo)
            n_elo.append(black_player.elo)
            elo_chg.append(black_player.elo - Decimal(black_old_elo))
            dat.append(game.game_date)
            ply.append(white_player.name)
            opp.append(black_player.name)
            o_elo.append(white_old_elo)
            n_elo.append(white_player.elo)
            elo_chg.append(white_player.elo - Decimal(white_old_elo))
            blk.append(black_player.name)
            wht.append(white_player.name)
            hdcp.append(game.handicap)
            gdat.append(game.game_date)
            rst.append(game.result)
            b_o_elo.append(black_old_elo)
            b_n_elo.append(black_player.elo)
            b_elo_chg.append(black_player.elo - Decimal(black_old_elo))
            w_o_elo.append(white_old_elo)
            w_n_elo.append(white_player.elo)
            w_elo_chg.append(white_player.elo - Decimal(white_old_elo))
            c+=2
        hist = {
            "Date": dat,
            "Player": ply,
            "Oppoent": opp,
            "Old Elo": o_elo,
            "New Elo": n_elo,
            "Elo Change": elo_chg
        }
        df1 = pd.DataFrame(hist)
        df1.to_csv(csv_out_path + 'history.csv', index=False)
        gme = {
            "Date": gdat,
            "Black": blk,
            "White": wht,
            "Result": rst,
            "Handicap": hdcp,
            "Black Old Elo": b_o_elo,
            "Black New Elo": b_n_elo,
            "Black Elo Change": b_elo_chg,
            "White Old Elo": w_o_elo,
            "White New Elo": w_n_elo,
            "White Elo Change": w_elo_chg
        }
        df3 = pd.DataFrame(gme)
        df3.to_csv(csv_out_path + 'game.csv', index=False)

        elo_sheet = book.create_sheet('Elo')
        elo_sheet.cell(row=1, column=1).value = 'Player'
        elo_sheet.cell(row=1, column=2).value = 'Elo'
        elo_sheet.cell(row=1, column=3).value = 'Total Games'
        elo_sheet.cell(row=1, column=4).value = 'Win'
        elo_sheet.cell(row=1, column=5).value = 'Lost'
        elo_sheet.cell(row=1, column=6).value = 'Win Rate'
        elo_sheet.cell(row=1, column=7).value = 'Status'
        rk = []
        play = []
        el = []
        ttl = []
        w = []
        l = []
        wr = []
        st = []
        c=2
        players = Player.objects.order_by('-elo')
        for player in players:
            latest_game = Game.objects.filter(Q(black=player) | Q(white=player)).order_by('-game_date').first()
            if latest_game.game_date + timedelta(days=180) < date.today():
                player.status = 'inactive'
            elif player.total_games <= 5:
                player.status = 'new'
            else:
                player.status = 'normal'
            player.winrate = str(round((player.win / player.total_games) * 10000) / 100) + '%'
            player.save()

            elo_sheet.cell(row=c, column=1).value = player.name
            elo_sheet.cell(row=c, column=2).value = player.elo
            elo_sheet.cell(row=c, column=3).value = player.total_games
            elo_sheet.cell(row=c, column=4).value = player.win
            elo_sheet.cell(row=c, column=5).value = player.lost
            elo_sheet.cell(row=c, column=6).value = player.winrate
            elo_sheet.cell(row=c, column=7).value = player.status
            play.append(player.name)
            el.append(player.elo)
            ttl.append(player.total_games)
            w.append(player.win)
            l.append(player.lost)
            wr.append(player.winrate)
            st.append(player.status)
            rk.append(c-1)
            c+=1
        
        elo = {
            "Rank": rk,
            "Player": play,
            "Elo": el,
            "Total Games": ttl,
            "Win": w,
            "Lost": l,
            "Win Rate": wr,
            "Status": st
        }
        df2 = pd.DataFrame(elo)
        df2.to_csv(csv_out_path + 'elo.csv', index=False)

        book.save(out_source)
        book.close()
    except:
        book.close()
        traceback.print_exc()